#!/usr/bin/python
import sys, os, re
import argparse
import yaml

import example_parser

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.getcwd())))
example_parser.REPO = REPO

TYPE = [
    "checklist",
    "estimation",
    "status",
]

COMPILER = [
    "armgcc",
    "atl",
    "iar",
    "kds",
    "mdk",
]

TARGET = [
    "D",
    "R",
]

class Utils():

    # insert column after selected column
    def insert_column(self, worksheet, column):
        if not (worksheet and column):
            return None
        _cells = worksheet._cells
        from copy import copy
        for cell in reversed(sorted(_cells)):
            if cell[1] < column:
                continue
            worksheet.cell(row = cell[0], column = cell[1]+1).value = _cells[(cell[0], cell[1])].value
            style = copy(worksheet.cell(row = cell[0], column = cell[1]).style)
            worksheet.cell(row = cell[0], column = cell[1]+1).style = style
            del _cells[(cell[0], cell[1])]
        import openpyxl
        new_merged = []
        for merged in reversed(sorted(worksheet._merged_cells)):
            cells = merged.split(':')
            new_cell = []
            for _cell in cells:
                _column = re.sub(r'[0-9]*', '', _cell)
                _row = re.sub(r'[A-Za-z]*', '', _cell)
                _column = openpyxl.cell.cell.column_index_from_string(_column)
                _row = int(_row)
                if _column >= column:
                    _column += 1
                    _cell = "".join([openpyxl.cell.get_column_letter(_column), '%d' % _row])
                new_cell.append(_cell)
            new_merged.append(":".join(new_cell))
        worksheet._merged_cells = new_merged[:]

    def insert_row(self, worksheet, row):
        if not (worksheet and row):
            return None
        _cells = worksheet._cells
        from copy import copy
        for cell in reversed(sorted(_cells)):
            if cell[0] < row:
                continue
            worksheet.cell(row = cell[0]+1, column = cell[1]).value = _cells[(cell[0], cell[1])].value
            style = copy(worksheet.cell(row = cell[0], column = cell[1]).style)
            worksheet.cell(row = cell[0]+1, column = cell[1]).style = style
            del _cells[(cell[0], cell[1])]
        import openpyxl
        new_merged = []
        for merged in reversed(sorted(worksheet._merged_cells)):
            cells = merged.split(':')
            new_cell = []
            for _cell in cells:
                _column = re.sub(r'[0-9]*', '', _cell)
                _row = re.sub(r'[A-Za-z]*', '', _cell)
                _column = openpyxl.cell.cell.column_index_from_string(_column)
                _row = int(_row)
                if _row >= row:
                    _row += 1
                    _cell = "".join([openpyxl.cell.get_column_letter(_column), '%d' % _row])
                new_cell.append(_cell)
            new_merged.append(":".join(new_cell))
        worksheet._merged_cells = new_merged[:]

class Reporter():

    # initizialize font format
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    target_font = Font(name='Arial', size=10, bold=True)
    result_font = Font(name='Arial', size=10, bold=False)
    target1_color = PatternFill(start_color='D9E8EF', end_color='D9E8EF', fill_type='solid')
    target2_color = PatternFill(start_color='A4C1CF', end_color='A4C1CF', fill_type='solid')
    target3_color = PatternFill(start_color='CEE3A1', end_color='CEE3A1', fill_type='solid')
    confirm_color = PatternFill(start_color='AF7A3C', end_color='AF7A3C', fill_type='solid')
    update_color = PatternFill(start_color='AB3A42', end_color='AB3A42', fill_type='solid')
    na_color = PatternFill(start_color='C1C1C1', end_color='C1C1C1', fill_type='solid')
    avail_color = PatternFill(start_color='BDE8B6', end_color='BDE8B6', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')

    def __init__(self, argv):
        """
        """
        parser = argparse.ArgumentParser(description="SDK-Reporter")
        parser.add_argument('-type', type=str, help="Report type: estimation|status|checklist", choices = TYPE)
        parser.add_argument('-soc', nargs='+', type=str, help="List of soc, seperated by space")
        parser.add_argument('-report', required = False, type=str, help="Name of the Report file")
        args = parser.parse_args(argv)
        if not args.type in TYPE:
            parser.print_help()
            exit()
        if args.type == "estimation":
            # configs = self.parse_config(args.soc)
            self.process_estimation(config = args.soc)
        if args.type == "status":
            _report = args.report
            if args.soc:
                self.process_status(_report=_report, config=args.soc)
            else:
                self.process_status(_report=_report)
        if args.type == "checklist":
            # configs = self.parse_config(args.soc)
            _report = args.report
            self.process_checklist(_report = _report, config=args.soc)

    def parse_driver_IP_ver(self, socs):
        """
            parse <soc>-feature and meta file to get feature
            structure:
            SOC:
                board:
                - BOARD
                ...
                PART_NAME:
                    driver: ver
                    ...
        """
        from meta_parser import MetaParser
        result = {}
        all_boards = os.listdir(os.path.join(REPO, 'boards'))
        for soc in socs:
            meta =  MetaParser()
            meta.meta_get_device_part_num(soc)
            parts = meta.part_num_list
            result[soc] = {"board":[],}
            for part in parts:
                meta.meta_compare_to_get_drivers_dic(soc, part)
                result[soc][part] = {}
                for driver in meta.ipdriver_dic:
                    result[soc][part][driver] = meta.ipdriver_dic[driver].split(" ")
                for board in meta.boardtype_list:
                    board = board.replace('-', '').lower()
                    for r_board in all_boards:
                        # Calculate to detect the truest board name
                        if (board in r_board) and (not r_board in result[soc]["board"]):
                            result[soc]["board"].append(r_board)
        return result

    def parse_soc_driver(self, socs):
        result = {}
        for soc in socs:
            result[soc] = []
            lib_driver = os.path.join(REPO, "bin\\generator\\records\\ksdk\\lib\\lib_sdk", soc, "drivers.yml")
            if not soc in os.listdir(os.path.join(REPO, "bin\\generator\\records\\ksdk\\lib\\lib_sdk")):
                continue
            with open(lib_driver, 'r') as stream:
                load_list = yaml.load(stream)["__load__"]
                for loaded in load_list:
                    with open(os.path.join(REPO, 'bin\\generator\\records', loaded.replace('/','\\')), 'r') as _stream:
                        drv_record = yaml.load(_stream)
                        try:
                            for module in drv_record["__common__"]["modules"].keys():
                                result[soc].append(module)
                        except:
                            continue
        return result

    def get_feature(self, socs):
        from feature_parser import FeatureParser
        featureParser =  FeatureParser()
        result = {}
        for soc in socs:
            result[soc] = []
            for part in socs[soc]:
                if 'board' == part:
                    continue
                featureParser.get_featurelist(soc, part)
                feature_driver = featureParser.transfer_features2drivers()
                for feature in feature_driver:
                    if feature in result[soc]:
                            continue
                    result[soc].append(feature)
        return result

    def compare(self, _ver_a, _ver_b):
        _a = float(_ver_a[:_ver_a.index(".")+1] + _ver_a[_ver_a.index("."):].replace(".", ""))
        _b = float(_ver_b[:_ver_b.index(".")+1] + _ver_b[_ver_b.index("."):].replace(".", ""))
        return _b - _a

    def make_acronym_sheet(self, acronym_sheet):
        acronym_sheet.cell(row = 1, column = 1).value = 'A'
        acronym_sheet.cell(row = 1, column = 1).fill = self.avail_color
        acronym_sheet.cell(row = 1, column = 1).border = self.border
        acronym_sheet.cell(row = 1, column = 2).value = 'The project is required in generator'
        acronym_sheet.cell(row = 1, column = 2).border = self.border

        acronym_sheet.cell(row = 2, column = 1).value = 'n/a'
        acronym_sheet.cell(row = 2, column = 1).fill = self.na_color
        acronym_sheet.cell(row = 2, column = 1).border = self.border
        acronym_sheet.cell(row = 2, column = 2).value = 'The project is not required in generator'
        acronym_sheet.cell(row = 2, column = 2).border = self.border

        acronym_sheet.cell(row = 3, column = 1).value = 'FB'
        acronym_sheet.cell(row = 3, column = 1).fill = self.confirm_color
        acronym_sheet.cell(row = 3, column = 1).border = self.border
        acronym_sheet.cell(row = 3, column = 2).value = 'The project build failure'
        acronym_sheet.cell(row = 3, column = 2).border = self.border

        acronym_sheet.cell(row = 4, column = 1).value = 'FR'
        acronym_sheet.cell(row = 4, column = 1).fill = self.confirm_color
        acronym_sheet.cell(row = 4, column = 1).border = self.border
        acronym_sheet.cell(row = 4, column = 2).value = 'The project run failure'
        acronym_sheet.cell(row = 4, column = 2).border = self.border

        acronym_sheet.cell(row = 5, column = 1).value = 'Confirm'
        acronym_sheet.cell(row = 5, column = 1).fill = self.confirm_color
        acronym_sheet.cell(row = 5, column = 1).border = self.border
        acronym_sheet.cell(row = 5, column = 2).value = 'Developer must be check IP of this driver'
        acronym_sheet.cell(row = 5, column = 2).border = self.border

        acronym_sheet.cell(row = 6, column = 1).value = 'On-going'
        acronym_sheet.cell(row = 6, column = 1).fill = self.update_color
        acronym_sheet.cell(row = 6, column = 1).border = self.border
        acronym_sheet.cell(row = 6, column = 2).value = 'Driver which must be updated source code'
        acronym_sheet.cell(row = 6, column = 2).border = self.border

        acronym_sheet.cell(row = 7, column = 1).value = 'uncertain'
        acronym_sheet.cell(row = 7, column = 1).fill = self.update_color
        acronym_sheet.cell(row = 7, column = 1).border = self.border
        acronym_sheet.cell(row = 7, column = 2).value = 'Define of module in feature header file is not certain with IP block in meta file'
        acronym_sheet.cell(row = 7, column = 2).border = self.border

        acronym_sheet.cell(row = 8, column = 1).value = 'unsupport'
        acronym_sheet.cell(row = 8, column = 1).fill = self.update_color
        acronym_sheet.cell(row = 8, column = 1).border = self.border
        acronym_sheet.cell(row = 8, column = 2).value = 'Not found feature IP'
        acronym_sheet.cell(row = 8, column = 2).border = self.border

        acronym_sheet.column_dimensions['B'].width = 45

    def process_estimation(self, config, _report = None):
        sheet_names = [
            'acronym',
            # 'new_IP',
            'driver',
            'demo_apps',
            'driver_examples',
            'usb',
        ]

        demo_apps_require, driver_examples_require, usb_require = example_parser.get_examples_module()

        workbook = Workbook()
        ws = workbook.active
        ws.title = sheet_names[0]

        for missing in sheet_names[1:]:
            workbook.create_sheet(title=missing)
        for _sheet in sheet_names:
            globals()["%s_sheet" % _sheet] = workbook.get_sheet_by_name(_sheet)

        socs = self.parse_driver_IP_ver(config)
        # driver_require = example_parser.get_driver_require()
        available_feature = self.get_feature(socs)
        # new_IP_sheet.cell(row = 1, column = 1).value = "IP"
        # new_IP_sheet.cell(row = 1, column = 1).font = self.target_font
        # new_IP_sheet.cell(row = 1, column = 1).fill = self.target1_color
        # new_IP_sheet.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column  = 1)
        # index = 2
        # for soc in sorted(socs):
        #     new_IP_sheet.cell(row = 1, column = index).value = soc
        #     for part in socs[soc]:
        #         if part == 'board':
        #             continue
        #         new_IP_sheet.cell(row = 2, column = index).value = part
        #         new_IP_sheet.cell(row = 2, column = index).fill = self.target2_color
        #         cur_column = re.sub(r'[0-9]', '', new_IP_sheet.cell(column=index, row=2).coordinate)
        #         new_IP_sheet.column_dimensions[cur_column].width = len(part)*1.3
        #         for driver in sorted(socs[soc][part]):
        #             count = 3
        #             while ((new_IP_sheet.cell(column=1, row=count).value != driver)
        #                     and (new_IP_sheet.cell(column=1, row=count).value)):
        #                 count += 1
        #             if not driver in driver_require:
        #                 continue
        #             if socs[soc][part][driver] in driver_require[driver] or socs[soc][part][driver] == "None":
        #                 continue
        #             new_IP_sheet.cell(column=index, row=count).font = self.result_font
        #             new_IP_sheet.cell(column=index, row=count).value = socs[soc][part][driver]
        #             if not new_IP_sheet.cell(column=1, row=count).value:
        #                 new_IP_sheet.cell(column=1, row=count).value = driver
        #                 new_IP_sheet.cell(column=1, row=count).font = self.target_font
        #                 new_IP_sheet.cell(column=1, row=count).fill = self.target3_color
        #         index += 1
        # max_size = 10
        # count = 3
        # while new_IP_sheet.cell(column=1, row=count).value:
        #     _value = len(new_IP_sheet.cell(column=1, row=count).value)
        #     if _value > max_size:
        #         max_size = _value
        #     count += 1
        # soc_size = 3
        # while new_IP_sheet.cell(column=soc_size, row=2).value:
        #     soc_size += 1
        # new_IP_sheet.column_dimensions['A'].width = max_size*1.3
        # # new_IP_sheet.column_dimensions['A'].hidden = False
        # for _y in range(1,soc_size):
        #     for _x in range(1,count):
        #         new_IP_sheet.cell(column=_y, row=_x).border = self.border
        #         new_IP_sheet.cell(column=_y, row=_x).alignment = self.alignment
        #         if not new_IP_sheet.cell(column=_y, row=_x).value:
        #             new_IP_sheet.cell(column=_y, row=_x).fill = self.na_color
        # start_merge = 2
        # while new_IP_sheet.cell(column=start_merge, row=2).value:
        #     count = 1
        #     while ((not new_IP_sheet.cell(column=start_merge+count, row=1).value)
        #         and new_IP_sheet.cell(column=start_merge+count, row=2).value):
        #         count += 1
        #     if count > 1:
        #         new_IP_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
        #         new_IP_sheet.cell(column=start_merge,row=1).font = self.target_font
        #         new_IP_sheet.cell(column=start_merge,row=1).fill = self.target1_color
        #     start_merge += count

        #
        driver_sheet.cell(row = 1, column = 1).value = "Driver"
        driver_sheet.cell(row = 1, column = 1).font = self.target_font
        driver_sheet.cell(row = 1, column = 1).fill = self.target1_color
        for index, soc in enumerate(sorted(socs)):
            driver_sheet.cell(row = 1, column = index + 2).value = soc
            driver_sheet.cell(row = 1, column = index + 2).font = self.target_font
            driver_sheet.cell(row = 1, column = index + 2).fill = self.target1_color
            for feature in sorted(available_feature[soc]):
                count = 2
                while ((driver_sheet.cell(column=1, row=count).value != feature)
                        and (driver_sheet.cell(column=1, row=count).value)):
                    count += 1
                _for_check = False
                _uncertain_check = False
                _unsupport_check = False
                for part in socs[soc]:
                    if part == 'board':
                        continue
                    if not feature in socs[soc][part]:
                        _for_check = True
                        continue
                    elif socs[soc][part][feature][0] == 'uncertain':
                        _uncertain_check = True
                        continue
                    elif socs[soc][part][feature][1] == 'unsupport':
                        _unsupport_check = True
                        continue
                if _for_check:
                    driver_sheet.cell(column=index + 2, row=count).value = "on-going"
                    driver_sheet.cell(column=index + 2, row=count).fill = self.update_color
                elif _uncertain_check:
                    driver_sheet.cell(column=index + 2, row=count).value = "uncertain"
                    driver_sheet.cell(column=index + 2, row=count).fill = self.confirm_color
                elif _unsupport_check:
                    driver_sheet.cell(column=index + 2, row=count).value = "on-going"
                    driver_sheet.cell(column=index + 2, row=count).fill = self.update_color
                else:
                    driver_sheet.cell(column=index + 2, row=count).value = "available"
                    driver_sheet.cell(column=index + 2, row=count).fill = self.avail_color
                if not driver_sheet.cell(column=1, row=count).value:
                    driver_sheet.cell(column=1, row=count).value = feature
                    driver_sheet.cell(column=1, row=count).font = self.target_font
                    driver_sheet.cell(column=1, row=count).fill = self.target3_color
        max_size = 10
        count = 2
        while driver_sheet.cell(column=1, row=count).value:
            _value = len(driver_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        soc_size = 2
        while driver_sheet.cell(column=soc_size, row=1).value:
            soc_size += 1
        driver_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_size):
            for _x in range(1,count):
                driver_sheet.cell(column=_y, row=_x).border = self.border
                driver_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not driver_sheet.cell(column=_y, row=_x).value:
                    driver_sheet.cell(column=_y, row=_x).fill = self.na_color

        boards = {}
        for soc in socs:
            for board in socs[soc]['board']:
                boards[board] = soc
        demo_apps_sheet.cell(row = 1, column = 1).value = "Demo_apps"
        demo_apps_sheet.cell(row = 1, column = 1).font = self.target_font
        demo_apps_sheet.cell(row = 1, column = 1).fill = self.target1_color
        index = 2
        for index, board in enumerate(sorted(boards)):
            demo_apps_sheet.cell(row = 1, column = 2+index).value = board
            demo_apps_sheet.cell(row = 1, column = 2+index).font = self.target_font
            demo_apps_sheet.cell(row = 1, column = 2+index).fill = self.target1_color
            for demo in sorted(demo_apps_require):
                count = 2
                while (demo_apps_sheet.cell(column=1, row=count).value != demo) and (demo_apps_sheet.cell(column=1, row=count).value):
                    count += 1
                demo_apps_sheet.cell(column=index + 2, row=count).value = "on-going"
                for drv in demo_apps_require[demo]:
                    if drv == 'uart':
                        if not re.search(r'(lpuart)|(uart)|(lpsci)', ' '.join(available_feature[boards[board]])):
                            demo_apps_sheet.cell(column=index + 2, row=count).value = "not support"
                            break
                        continue
                    if not drv in ' '.join(available_feature[boards[board]]):
                        demo_apps_sheet.cell(column=index + 2, row=count).value = "not support"
                        break
                if demo_apps_sheet.cell(column=index + 2, row=count).value == "on-going":
                    demo_apps_sheet.cell(column=index + 2, row=count).fill = self.avail_color
                if not demo_apps_sheet.cell(column=1, row=count).value:
                    demo_apps_sheet.cell(column=1, row=count).value = demo
                    demo_apps_sheet.cell(column=1, row=count).font = self.target_font
                    demo_apps_sheet.cell(column=1, row=count).fill = self.target3_color
        max_size = 10
        count = 2
        while demo_apps_sheet.cell(column=1, row=count).value:
            _value = len(demo_apps_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        soc_size = 2
        while demo_apps_sheet.cell(column=soc_size, row=1).value:
            soc_size += 1
        demo_apps_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_size):
            for _x in range(1,count):
                demo_apps_sheet.cell(column=_y, row=_x).border = self.border
                demo_apps_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not demo_apps_sheet.cell(column=_y, row=_x).value:
                    demo_apps_sheet.cell(column=_y, row=_x).fill = self.na_color

        driver_examples_sheet.cell(row = 1, column = 1).value = "Driver_examples"
        driver_examples_sheet.cell(row = 1, column = 1).font = self.target_font
        driver_examples_sheet.cell(row = 1, column = 1).fill = self.target1_color
        index = 2
        for index, board in enumerate(sorted(boards)):
            driver_examples_sheet.cell(row = 1, column = 2+index).value = board
            driver_examples_sheet.cell(row = 1, column = 2+index).font = self.target_font
            driver_examples_sheet.cell(row = 1, column = 2+index).fill = self.target1_color
            for examples in sorted(driver_examples_require):
                count = 2
                while (driver_examples_sheet.cell(column=1, row=count).value != examples) and (driver_examples_sheet.cell(column=1, row=count).value):
                    count += 1
                driver_examples_sheet.cell(column=index + 2, row=count).value = "on-going"
                for drv in driver_examples_require[examples]:
                    if drv == 'uart':
                        if not re.search(r'(lpuart)|(uart)|(lpsci)', ' '.join(available_feature[boards[board]])):
                            driver_examples_sheet.cell(column=index + 2, row=count).value = "not support"
                            break
                        continue
                    if not drv in ' '.join(available_feature[boards[board]]):
                        driver_examples_sheet.cell(column=index + 2, row=count).value = "not support"
                        break
                if driver_examples_sheet.cell(column=index + 2, row=count).value == "on-going":
                    driver_examples_sheet.cell(column=index + 2, row=count).fill = self.avail_color
                if not driver_examples_sheet.cell(column=1, row=count).value:
                    driver_examples_sheet.cell(column=1, row=count).value = examples
                    driver_examples_sheet.cell(column=1, row=count).font = self.target_font
                    driver_examples_sheet.cell(column=1, row=count).fill = self.target3_color
        max_size = 10
        count = 2
        while driver_examples_sheet.cell(column=1, row=count).value:
            _value = len(driver_examples_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        soc_size = 2
        while driver_examples_sheet.cell(column=soc_size, row=1).value:
            soc_size += 1
        driver_examples_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_size):
            for _x in range(1,count):
                driver_examples_sheet.cell(column=_y, row=_x).border = self.border
                driver_examples_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not driver_examples_sheet.cell(column=_y, row=_x).value:
                    driver_examples_sheet.cell(column=_y, row=_x).fill = self.na_color

        usb_sheet.cell(row = 1, column = 1).value = "USB"
        usb_sheet.cell(row = 1, column = 1).font = self.target_font
        usb_sheet.cell(row = 1, column = 1).fill = self.target1_color
        index = 2
        for index, board in enumerate(sorted(boards)):
            usb_sheet.cell(row = 1, column = 2+index).value = board
            usb_sheet.cell(row = 1, column = 2+index).font = self.target_font
            usb_sheet.cell(row = 1, column = 2+index).fill = self.target1_color
            for usb in sorted(usb_require):
                count = 2
                while (usb_sheet.cell(column=1, row=count).value != usb) and (usb_sheet.cell(column=1, row=count).value):
                    count += 1
                usb_sheet.cell(column=index + 2, row=count).value = "on-going"
                for drv in usb_require[usb]:
                    if drv == 'uart':
                        if not re.search(r'(lpuart)|(uart)|(lpsci)', ' '.join(available_feature[boards[board]])):
                            usb_sheet.cell(column=index + 2, row=count).value = "not support"
                            break
                        continue
                    if not drv in ' '.join(available_feature[boards[board]]):
                        usb_sheet.cell(column=index + 2, row=count).value = "not support"
                        break
                if usb_sheet.cell(column=index + 2, row=count).value == "on-going":
                    usb_sheet.cell(column=index + 2, row=count).fill = self.avail_color
                if not usb_sheet.cell(column=1, row=count).value:
                    usb_sheet.cell(column=1, row=count).value = usb
                    usb_sheet.cell(column=1, row=count).font = self.target_font
                    usb_sheet.cell(column=1, row=count).fill = self.target3_color
        max_size = 10
        count = 2
        while usb_sheet.cell(column=1, row=count).value:
            _value = len(usb_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        soc_size = 2
        while usb_sheet.cell(column=soc_size, row=1).value:
            soc_size += 1
        usb_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_size):
            for _x in range(1,count):
                usb_sheet.cell(column=_y, row=_x).border = self.border
                usb_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not usb_sheet.cell(column=_y, row=_x).value:
                    usb_sheet.cell(column=_y, row=_x).fill = self.na_color

        self.make_acronym_sheet(acronym_sheet)

        if not _report:
            _report = "estimation.xlsx"
        workbook.save(_report)

        status_book = Workbook()
        ws = status_book.active
        ws.title = 'feature_version'
        ws.cell(row = 1, column = 1).value = "IP"
        ws.cell(row = 1, column = 1).font = self.target_font
        ws.cell(row = 1, column = 1).fill = self.target1_color
        ws.cell(row = 1, column = 1).border = self.border
        ws.cell(row = 2, column = 1).border = self.border
        ws.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column  = 1)
        index = 2
        for soc in sorted(socs):
            ws.cell(row = 1, column = index).value = soc
            ws.cell(row = 1, column = index).fill = self.target1_color
            ws.cell(row = 1, column = index).border = self.border
            _index = 0
            for part in socs[soc]:
                if part == 'board':
                    continue
                ws.cell(row = 2, column = index+_index).value = part
                ws.cell(row = 2, column = index+_index).fill = self.target2_color
                ws.cell(row = 1, column = index+_index).border = self.border
                cur_column = re.sub(r'[0-9]', '', ws.cell(column=index+_index, row=2).coordinate)
                ws.column_dimensions[cur_column].width = 45
                _index += 1
            ws.merge_cells(start_row = 1, start_column = index, end_row = 1, end_column  = index + _index - 1)
            index += _index
        status_book.save("status.xlsx")
        self.process_status(_report = "status.xlsx", config = config)

    def process_status(self, _report, config = None):
        # initizialize excel sheets
        sheet_names = [
            'acronym',
            'feature_version',
            'devices_driver',
            'demo_apps',
            'driver_examples',
            'usb',
        ]

        workbook = load_workbook(_report)
        avail_sheets = workbook.get_sheet_names()
        missings = list(set(sheet_names) - set(avail_sheets))

        for missing in missings:
            workbook.create_sheet(title=missing)

        for _sheet in sheet_names:
            globals()["%s_sheet" % _sheet] = workbook.get_sheet_by_name(_sheet)

        soc_index = 2
        part_index = 0
        if not config:
            config = []
            while feature_version_sheet.cell(row = 2, column = soc_index+part_index).value:
                soc = str(feature_version_sheet.cell(row = 1, column = soc_index).value)
                if feature_version_sheet.cell(row = 1, column = soc_index+part_index).value:
                    soc_index += part_index
                    part_index = 0
                    config.append(soc)
                part_index += 1
        socs = self.parse_driver_IP_ver(config)
        for index,soc in enumerate(sorted(config)):
            print "process device sheet: %s" % soc
            soc_index = 2
            while (feature_version_sheet.cell(column=soc_index, row=1).value != soc
                        and feature_version_sheet.cell(column=soc_index, row=2).value):
                soc_index += 1
            for _index, part in enumerate(sorted(socs[soc])):
                if part == 'board':
                    continue
                part_index = 0
                while ((feature_version_sheet.cell(column=soc_index + part_index, row=1).value == soc or
                            (not feature_version_sheet.cell(column=soc_index + part_index, row=1).value))
                        and (feature_version_sheet.cell(column=soc_index + part_index, row=2).value != part)
                        and (feature_version_sheet.cell(column=soc_index + part_index, row=2).value)):
                    part_index += 1
                if part in socs[soc]:
                    for __index,driver in enumerate(sorted(socs[soc][part])):
                        count = 3
                        while ((feature_version_sheet.cell(column=1, row=count).value != driver)
                                and (feature_version_sheet.cell(column=1, row=count).value)):
                            count += 1
                        feature_version_sheet.cell(column=soc_index+part_index, row=count).font = self.result_font
                        feature_version_sheet.cell(column=soc_index+part_index, row=count).fill = self.avail_color
                        if not feature_version_sheet.cell(column=soc_index+part_index, row=count).value or feature_version_sheet.cell(column=soc_index+part_index, row=count).value == socs[soc][part][driver][1]:
                            feature_version_sheet.cell(column=soc_index+part_index, row=count).value = socs[soc][part][driver][1]
                            if socs[soc][part][driver][1] == "unsupport":
                                feature_version_sheet.cell(column=soc_index+part_index, row=count).fill = self.update_color
                        elif feature_version_sheet.cell(column=soc_index+part_index, row=count).value != socs[soc][part][driver][1]:
                            feature_version_sheet.cell(column=soc_index+part_index, row=count).fill = self.confirm_color

                        if not feature_version_sheet.cell(column=1, row=count).value:
                            feature_version_sheet.cell(column=1, row=count).value = driver
                            feature_version_sheet.cell(column=1, row=count).font = self.target_font
                            feature_version_sheet.cell(column=1, row=count).fill = self.target3_color

        max_size = 0
        count = 3
        while feature_version_sheet.cell(column=1, row=count).value:
            _value = len(feature_version_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        soc_size = 3
        while feature_version_sheet.cell(column=soc_size, row=2).value:
            soc_size += 1
        feature_version_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_size):
            for _x in range(1,count):
                feature_version_sheet.cell(column=_y, row=_x).border = self.border
                feature_version_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not feature_version_sheet.cell(column=_y, row=_x).value:
                    feature_version_sheet.cell(column=_y, row=_x).fill = self.na_color
        start_merge = 2
        while feature_version_sheet.cell(column=start_merge, row=2).value:
            count = 1
            while ((not feature_version_sheet.cell(column=start_merge+count, row=1).value)
                and feature_version_sheet.cell(column=start_merge+count, row=2).value):
                count += 1
            if count > 1:
                feature_version_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
                feature_version_sheet.cell(column=start_merge,row=1).font = self.target_font
                feature_version_sheet.cell(column=start_merge,row=1).fill = self.target1_color
            start_merge += count

        _boards = []
        for soc in socs:
            if not socs[soc]["board"]:
                continue
            _boards += socs[soc]["board"]
        demos, examples, usbs = example_parser.parse_examples(_boards)
        cur_soc_driver = self.parse_soc_driver(config)
        driver_require = example_parser.get_driver_require()

        devices_driver_sheet.cell(column=1, row = 1).value = "Driver"
        devices_driver_sheet.cell(column=1, row = 1).font = self.target_font
        devices_driver_sheet.cell(column=1, row = 1).fill = self.target1_color
        soc_index = 2
        for index,soc in enumerate(sorted(config)):
            print "process drivers sheet: %s" % soc
            devices_driver_sheet.cell(column=soc_index, row=1).value = soc
            devices_driver_sheet.cell(column=soc_index, row=1).fill = self.target1_color
            devices_driver_sheet.cell(column=soc_index, row=1).font = self.target_font

            cur_column = re.sub(r'[0-9]', '', devices_driver_sheet.cell(column=soc_index, row=1).coordinate)
            devices_driver_sheet.column_dimensions[cur_column].width = len(soc)*1.3

            if not soc in cur_soc_driver:
                continue
            for drv in sorted(driver_require):
                count = 2
                while (devices_driver_sheet.cell(column=1, row=count).value != drv) and (devices_driver_sheet.cell(column=1, row=count).value):
                    count += 1
                devices_driver_sheet.cell(column=soc_index, row=count).font = self.result_font
                _for_check = False
                _uncertain_check = False
                _unsupport_check = False
                ips = []
                for part in socs[soc]:
                    if part == 'board':
                        continue
                    if not drv in socs[soc][part]:
                        _for_check = True
                        continue
                    elif socs[soc][part][drv][0] == 'uncertain':
                        _uncertain_check = True
                        continue
                    elif socs[soc][part][drv][1] == 'unsupport':
                        _unsupport_check = True
                        continue
                    ips.append(socs[soc][part][drv][1])
                if not devices_driver_sheet.cell(column=soc_index, row=count).value:
                    devices_driver_sheet.cell(column=soc_index, row=count).value = 'n/a'
                    if _uncertain_check:
                        devices_driver_sheet.cell(column=soc_index, row=count).value = 'uncertain'
                        devices_driver_sheet.cell(column=soc_index, row=count).fill = self.update_color
                    elif _unsupport_check:
                        devices_driver_sheet.cell(column=soc_index, row=count).value = 'on-going'
                        devices_driver_sheet.cell(column=soc_index, row=count).fill = self.update_color
                    elif not _for_check:
                        devices_driver_sheet.cell(column=soc_index, row=count).value = 'Available'
                        devices_driver_sheet.cell(column=soc_index, row=count).fill = self.avail_color
                elif devices_driver_sheet.cell(column=soc_index, row=count).value == "Available":
                    if _uncertain_check or _unsupport_check:
                        devices_driver_sheet.cell(column=soc_index, row=count).value = 'confirm'
                        devices_driver_sheet.cell(column=soc_index, row=count).fill = self.confirm_color
                elif _for_check:
                    del devices_driver_sheet._cells[(count, soc_index)]
                if not devices_driver_sheet.cell(column=1, row=count).value:
                    devices_driver_sheet.cell(column=1, row=count).value = drv
                    devices_driver_sheet.cell(column=1, row=count).font = self.target_font
                    devices_driver_sheet.cell(column=1, row=count).fill = self.target3_color

            soc_index+=1
        max_size = len(devices_driver_sheet.cell(column=1, row=1).value)
        count = 2
        while devices_driver_sheet.cell(column=1, row=count).value:
            _value = len(devices_driver_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        devices_driver_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_index):
            for _x in range(1,count):
                devices_driver_sheet.cell(column=_y, row=_x).border = self.border
                devices_driver_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if devices_driver_sheet.cell(column=_y, row=_x).value in ['n/a', None]:
                    devices_driver_sheet.cell(column=_y, row=_x).fill = self.na_color

        if "demo_apps" in missings:
            demo_apps_sheet.cell(column=1, row = 1).value = "Demo_apps"
            demo_apps_sheet.merge_cells(start_row=1,start_column=1,end_row=3,end_column=1)
            demo_apps_sheet.cell(column=1, row = 1).font = self.target_font
            demo_apps_sheet.cell(column=1, row = 1).fill = self.target1_color
        for index,board in enumerate(sorted(demos)):
            if not demos[board]:
                continue
            if not demos[board]["soc"] in config:
                continue
            print "process demo_apps sheet: %s" % board

            board_index = 2
            while demo_apps_sheet.cell(column=board_index, row=3).value and (demo_apps_sheet.cell(column=board_index, row=1).value != board):
                board_index += 1

            if not demo_apps_sheet.cell(column=board_index, row=1).value:
                demo_apps_sheet.cell(column=board_index, row=1).value = board
                demo_apps_sheet.cell(column=board_index, row=1).fill = self.target1_color
                demo_apps_sheet.cell(column=board_index, row=1).font = self.target_font
            for _index, compiler in enumerate(sorted(COMPILER)):
                demo_apps_sheet.cell(column=board_index+_index*2, row=2).value = compiler
                demo_apps_sheet.cell(column=board_index+_index*2, row=2).font = self.result_font
                demo_apps_sheet.cell(column=board_index+_index*2, row=2).fill = self.target2_color
                demo_apps_sheet.merge_cells(start_row=2,start_column=board_index+_index*2,end_row=2,end_column=board_index+_index*2+1)
                for __index, target in enumerate(TARGET):
                    demo_apps_sheet.cell(column=board_index+_index*2+__index, row=3).value = target
                    demo_apps_sheet.cell(column=board_index+_index*2+__index, row=3).font = self.result_font
                    demo_apps_sheet.cell(column=board_index+_index*2+__index, row=3).fill = self.target2_color
                    cur_column = re.sub(r'[0-9]', '', demo_apps_sheet.cell(column=board_index+_index*2+__index, row=2).coordinate)
                    demo_apps_sheet.column_dimensions[cur_column].width = 4

                for __index, demo in enumerate(sorted(demos[board])):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (demo_apps_sheet.cell(column=1, row=count).value != demo) and (demo_apps_sheet.cell(column=1, row=count).value):
                        count += 1
                    demo_apps_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                    if not demo_apps_sheet.cell(column=board_index+_index*2, row=count).value:
                        if compiler in demos[board][demo]:
                            if demos[board][demo][compiler] in ['A', 'Do']:
                                demo_apps_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                                demo_apps_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', demo_apps_sheet.cell(column=board_index+_index*2, row=count).value):
                            if not compiler in demos[board][demo]:
                                del demo_apps_sheet._cells[(count,board_index+_index*2)]
                            elif demos[board][demo][compiler] in ['Ro', None]:
                                del demo_apps_sheet._cells[(count,board_index+_index*2)]
                            else:
                                demo_apps_sheet.cell(column=board_index+_index*2, row=count).value = "check!"
                                demo_apps_sheet.cell(column=board_index+_index*2, row=count).fill = self.confirm_color
                    # demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.na_color
                    demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
                    if not demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).value:
                        if compiler in demos[board][demo]:
                            if demos[board][demo][compiler] in ['A', 'Ro']:
                                demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                                demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).value):
                            if not compiler in demos[board][demo]:
                                del demo_apps_sheet._cells[(count,board_index+_index*2+1)]
                            elif demos[board][demo][compiler] in ['Do', None]:
                                del demo_apps_sheet._cells[(count,board_index+_index*2+1)]
                            else:
                                demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).value = "check!"
                                demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.confirm_color
                    if not demo_apps_sheet.cell(column=1, row=count).value:
                        demo_apps_sheet.cell(column=1, row=count).value = demo
                        demo_apps_sheet.cell(column=1, row=count).font = self.target_font
                        demo_apps_sheet.cell(column=1, row=count).fill = self.target3_color
                    demo_apps_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                    demo_apps_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
        max_size = len(demo_apps_sheet.cell(column=1, row=1).value)
        count = 4
        while demo_apps_sheet.cell(column=1, row=count).value:
            _value = len(demo_apps_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        demo_apps_sheet.column_dimensions['A'].width = max_size*1.3
        board_index = 2
        while demo_apps_sheet.cell(row=3, column=board_index).value:
            board_index += 10
        for _y in range(1,board_index):
            for _x in range(1,count):
                demo_apps_sheet.cell(column=_y, row=_x).border = self.border
                demo_apps_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not demo_apps_sheet.cell(column=_y, row=_x).value:
                    demo_apps_sheet.cell(column=_y, row=_x).fill = self.na_color
        start_merge = 2
        while demo_apps_sheet.cell(column=start_merge, row=3).value:
            count = 1
            while ((not demo_apps_sheet.cell(column=start_merge+count, row=1).value)
                and demo_apps_sheet.cell(column=start_merge+count, row=3).value):
                count += 1
            if count > 1:
                demo_apps_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
                demo_apps_sheet.cell(column=start_merge,row=1).font = self.target_font
                demo_apps_sheet.cell(column=start_merge,row=1).fill = self.target1_color
            start_merge += count

        if 'driver_examples' in missings:
            driver_examples_sheet.cell(column=1, row = 1).value = "Driver_examples"
            driver_examples_sheet.merge_cells(start_row=1,start_column=1,end_row=3,end_column=1)
            driver_examples_sheet.cell(column=1, row = 1).font = self.target_font
            driver_examples_sheet.cell(column=1, row = 1).fill = self.target1_color
        for index,board in enumerate(sorted(examples)):
            if not examples[board]:
                continue
            if not examples[board]["soc"] in config:
                continue
            print "process driver_examples sheet: %s" % board

            board_index = 2
            while driver_examples_sheet.cell(column=board_index, row=3).value and (driver_examples_sheet.cell(column=board_index, row=1).value != board):
                board_index += 1

            if not driver_examples_sheet.cell(column=board_index, row=1).value:
                driver_examples_sheet.cell(column=board_index, row=1).value = board
                driver_examples_sheet.cell(column=board_index, row=1).fill = self.target1_color
                driver_examples_sheet.cell(column=board_index, row=1).font = self.target_font
            for _index, compiler in enumerate(sorted(COMPILER)):
                driver_examples_sheet.cell(column=board_index+_index*2, row=2).value = compiler
                driver_examples_sheet.cell(column=board_index+_index*2, row=2).font = self.result_font
                driver_examples_sheet.cell(column=board_index+_index*2, row=2).fill = self.target2_color
                driver_examples_sheet.merge_cells(start_row=2,start_column=board_index+_index*2,end_row=2,end_column=board_index+_index*2+1)
                for __index, target in enumerate(TARGET):
                    driver_examples_sheet.cell(column=board_index+_index*2+__index, row=3).value = target
                    driver_examples_sheet.cell(column=board_index+_index*2+__index, row=3).font = self.result_font
                    driver_examples_sheet.cell(column=board_index+_index*2+__index, row=3).fill = self.target2_color
                    cur_column = re.sub(r'[0-9]', '', driver_examples_sheet.cell(column=board_index+_index*2+__index, row=2).coordinate)
                    driver_examples_sheet.column_dimensions[cur_column].width = 4

                for __index, demo in enumerate(sorted(examples[board])):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (driver_examples_sheet.cell(column=1, row=count).value != demo) and (driver_examples_sheet.cell(column=1, row=count).value):
                        count += 1
                    driver_examples_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                    if not driver_examples_sheet.cell(column=board_index+_index*2, row=count).value:
                        if compiler in examples[board][demo]:
                            if examples[board][demo][compiler] in ['A', 'Do']:
                                driver_examples_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                                driver_examples_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', driver_examples_sheet.cell(column=board_index+_index*2, row=count).value):
                            if not compiler in examples[board][demo]:
                                del driver_examples_sheet._cells[(count,board_index+_index*2)]
                            elif examples[board][demo][compiler] in ['Ro', None]:
                                del driver_examples_sheet._cells[(count,board_index+_index*2)]
                            else:
                                driver_examples_sheet.cell(column=board_index+_index*2, row=count).value = "check!"
                                driver_examples_sheet.cell(column=board_index+_index*2, row=count).fill = self.confirm_color
                    driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
                    if not driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).value:
                        if compiler in examples[board][demo]:
                            if examples[board][demo][compiler] in ['A', 'Ro']:
                                driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                                driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).value):
                            if not compiler in examples[board][demo]:
                                del driver_examples_sheet._cells[(count,board_index+_index*2+1)]
                            elif examples[board][demo][compiler] in ['Do', None]:
                                del driver_examples_sheet._cells[(count,board_index+_index*2+1)]
                            else:
                                driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).value = "check!"
                                driver_examples_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.confirm_color
                    if not driver_examples_sheet.cell(column=1, row=count).value:
                        driver_examples_sheet.cell(column=1, row=count).value = demo
                        driver_examples_sheet.cell(column=1, row=count).font = self.target_font
                        driver_examples_sheet.cell(column=1, row=count).fill = self.target3_color
                    driver_examples_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
        max_size = len(driver_examples_sheet.cell(column=1, row=1).value)
        count = 4
        while driver_examples_sheet.cell(column=1, row=count).value:
            _value = len(driver_examples_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        driver_examples_sheet.column_dimensions['A'].width = max_size*1.3
        board_index = 2
        while driver_examples_sheet.cell(row=3, column=board_index).value:
            board_index += 10
        for _y in range(1,board_index):
            for _x in range(1,count):
                driver_examples_sheet.cell(column=_y, row=_x).border = self.border
                driver_examples_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not driver_examples_sheet.cell(column=_y, row=_x).value:
                    driver_examples_sheet.cell(column=_y, row=_x).fill = self.na_color
        start_merge = 2
        while driver_examples_sheet.cell(column=start_merge, row=3).value:
            count = 1
            while ((not driver_examples_sheet.cell(column=start_merge+count, row=1).value)
                and driver_examples_sheet.cell(column=start_merge+count, row=3).value):
                count += 1
            if count > 1:
                driver_examples_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
                driver_examples_sheet.cell(column=start_merge,row=1).font = self.target_font
                driver_examples_sheet.cell(column=start_merge,row=1).fill = self.target1_color
            start_merge += count

        if 'usb' in missings:
            usb_sheet.cell(column=1, row = 1).value = "USB examples"
            usb_sheet.merge_cells(start_row=1,start_column=1,end_row=3,end_column=1)
            usb_sheet.cell(column=1, row = 1).font = self.target_font
            usb_sheet.cell(column=1, row = 1).fill = self.target1_color
        for index,board in enumerate(sorted(usbs)):
            if not usbs[board]:
                continue
            if not usbs[board]["soc"] in config:
                continue
            print "process usb sheet: %s" % board

            board_index = 2
            while usb_sheet.cell(column=board_index, row=3).value and (usb_sheet.cell(column=board_index, row=1).value != board):
                board_index += 1

            if not usb_sheet.cell(column=board_index, row=1).value:
                usb_sheet.cell(column=board_index, row=1).value = board
                usb_sheet.cell(column=board_index, row=1).fill = self.target1_color
                usb_sheet.cell(column=board_index, row=1).font = self.target_font
            for _index, compiler in enumerate(sorted(COMPILER)):
                usb_sheet.cell(column=board_index+_index*2, row=2).value = compiler
                usb_sheet.cell(column=board_index+_index*2, row=2).font = self.result_font
                usb_sheet.cell(column=board_index+_index*2, row=2).fill = self.target2_color
                usb_sheet.merge_cells(start_row=2,start_column=board_index+_index*2,end_row=2,end_column=board_index+_index*2+1)

                for __index, target in enumerate(TARGET):
                    usb_sheet.cell(column=board_index+_index*2+__index, row=3).value = target
                    usb_sheet.cell(column=board_index+_index*2+__index, row=3).font = self.result_font
                    usb_sheet.cell(column=board_index+_index*2+__index, row=3).fill = self.target2_color
                    cur_column = re.sub(r'[0-9]', '', usb_sheet.cell(column=board_index+_index*2+__index, row=2).coordinate)
                    usb_sheet.column_dimensions[cur_column].width = 4

                for __index, demo in enumerate(sorted(usbs[board])):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (usb_sheet.cell(column=1, row=count).value != demo) and (usb_sheet.cell(column=1, row=count).value):
                        count += 1
                    usb_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                    if not usb_sheet.cell(column=board_index+_index*2, row=count).value:
                        if compiler in usbs[board][demo]:
                            if usbs[board][demo][compiler] in ['A', 'Do']:
                                usb_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                                usb_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', usb_sheet.cell(column=board_index+_index*2, row=count).value):
                            if not compiler in usbs[board][demo]:
                                del usb_sheet._cells[(count,board_index+_index*2)]
                            elif usbs[board][demo][compiler] in ['Ro', None]:
                                del usb_sheet._cells[(count,board_index+_index*2)]
                            else:
                                usb_sheet.cell(column=board_index+_index*2, row=count).value = "check!"
                                usb_sheet.cell(column=board_index+_index*2, row=count).fill = self.confirm_color
                    usb_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
                    if not usb_sheet.cell(column=board_index+_index*2+1, row=count).value:
                        if compiler in usbs[board][demo]:
                            if usbs[board][demo][compiler] in ['A', 'Ro']:
                                usb_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                                usb_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                    else:
                        if re.search(r'[fF]', usb_sheet.cell(column=board_index+_index*2+1, row=count).value):
                            if not compiler in usbs[board][demo]:
                                del usb_sheet._cells[(count,board_index+_index*2+1)]
                            elif usbs[board][demo][compiler] in ['Do', None]:
                                del usb_sheet._cells[(count,board_index+_index*2+1)]
                            else:
                                usb_sheet.cell(column=board_index+_index*2+1, row=count).value = "check!"
                                usb_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.confirm_color
                    if not usb_sheet.cell(column=1, row=count).value:
                        usb_sheet.cell(column=1, row=count).value = demo
                        usb_sheet.cell(column=1, row=count).font = self.target_font
                        usb_sheet.cell(column=1, row=count).fill = self.target3_color
        max_size = len(usb_sheet.cell(column=1, row=1).value)
        count = 4
        while usb_sheet.cell(column=1, row=count).value:
            _value = len(usb_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        usb_sheet.column_dimensions['A'].width = max_size*1.3
        board_index = 2
        while usb_sheet.cell(row=3, column=board_index).value:
            board_index += 1
        for _y in range(1,board_index):
            for _x in range(1,count):
                usb_sheet.cell(column=_y, row=_x).border = self.border
                usb_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not usb_sheet.cell(column=_y, row=_x).value:
                    usb_sheet.cell(column=_y, row=_x).fill = self.na_color
        start_merge = 2
        while usb_sheet.cell(column=start_merge, row=2).value:
            count = 1
            while ((not usb_sheet.cell(column=start_merge+count, row=1).value)
                and usb_sheet.cell(column=start_merge+count, row=3).value):
                count += 1
            if count > 1:
                usb_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
                usb_sheet.cell(column=start_merge,row=1).font = self.target_font
                usb_sheet.cell(column=start_merge,row=1).fill = self.target1_color
            start_merge += count

        if "acronym" in missings:
            self.make_acronym_sheet(acronym_sheet)

        # Order work sheet
        worksheets = []
        for _sheet in sheet_names:
            worksheets.append(workbook.get_sheet_by_name(_sheet))
        workbook._sheets = worksheets

        #save workbook
        if not _report:
            _report = "status.xlsx"
        workbook.save(filename=_report)

    def process_checklist(self, config, _report):
        # checking method aguments
        socs = self.parse_driver_IP_ver(config)
        _boards = []
        for soc in socs:
            if not socs[soc]["board"]:
                continue
            _boards += socs[soc]["board"]
        _demos, _examples, _usbs = example_parser.parse_examples(_boards)
        examples = _examples
        demos = _demos
        usbs = _usbs
        soc_driver = self.parse_soc_driver(socs)

        # initizialize excel sheets
        sheet_names = [
            'information',
            'drivers',
            'examples_status',
        ]

        workbook = Workbook()
        _sheet = workbook.active
        _sheet.title = sheet_names[0]
        missings = sheet_names[1:]
        for missing in missings:
            workbook.create_sheet(title=missing)
        for _sheet in sheet_names:
            globals()["%s_sheet" % _sheet] = workbook.get_sheet_by_name(_sheet)

        information_sheet.cell(column=1, row = 1).value = "Boards:"
        next_table_index = 2
        information_sheet.cell(column=1, row = next_table_index).value = "SOC"
        information_sheet.cell(column=1, row = next_table_index).font = self.target_font
        information_sheet.cell(column=1, row = next_table_index).fill = self.target1_color
        information_sheet.cell(column=2, row = next_table_index).value = "Board"
        information_sheet.cell(column=2, row = next_table_index).font = self.target_font
        information_sheet.cell(column=2, row = next_table_index).fill = self.target1_color
        buff = 3
        for index, soc in enumerate(sorted(config)):
            information_sheet.cell(column=1, row = buff).value = soc
            next_table_index = buff
            for _index, board in enumerate(sorted(socs[soc]["board"])):
                information_sheet.cell(column=2, row = buff).value = board
                buff += 1
            if (buff-1) > next_table_index:
                information_sheet.merge_cells(start_column=1, start_row=next_table_index, end_column=1, end_row=buff-1)
            next_table_index = buff

        next_table_index = sorted(information_sheet._cells)[-1][0] + 1
        information_sheet.cell(column=1, row = next_table_index).value = "Part number:"
        next_table_index += 1
        information_sheet.cell(column=1, row = next_table_index).value = "SOC"
        information_sheet.cell(column=1, row = next_table_index).font = self.target_font
        information_sheet.cell(column=1, row = next_table_index).fill = self.target1_color
        information_sheet.cell(column=2, row = next_table_index).value = "Part number"
        information_sheet.cell(column=2, row = next_table_index).font = self.target_font
        information_sheet.cell(column=2, row = next_table_index).fill = self.target1_color
        # information_sheet.cell(column=3, row = next_table_index).value = "Ram size"
        # information_sheet.cell(column=3, row = next_table_index).font = self.target_font
        # information_sheet.cell(column=3, row = next_table_index).fill = self.target1_color
        # information_sheet.cell(column=4, row = next_table_index).value = "Flash size"
        # information_sheet.cell(column=4, row = next_table_index).font = self.target_font
        # information_sheet.cell(column=4, row = next_table_index).fill = self.target1_color
        buff = 1 + next_table_index
        for index, soc in enumerate(sorted(config)):
            information_sheet.cell(column=1, row = buff).value = soc
            next_table_index = buff
            for _index, part in enumerate(sorted(socs[soc])):
                if part == "board":
                    continue
                information_sheet.cell(column=2, row = buff).value = part
                buff += 1
            if (buff-1) > next_table_index:
                information_sheet.merge_cells(start_column=1, start_row=next_table_index, end_column=1, end_row=buff-1)
            next_table_index = buff

        release_info_file = os.path.join(REPO, "bin\\generator\\records\\ksdk\\options\\common\\releaseinfo.yml")
        with open(release_info_file, 'rb') as stream:
            release_info = yaml.load(stream)

        next_table_index = sorted(information_sheet._cells)[-1][0] + 1
        information_sheet.cell(column=1, row = next_table_index).value = "Toolchains:"
        next_table_index += 1
        for compiler in sorted(release_info["__common__"]["configuration"]["releaseinfo"]["toolchain"]):
            information_sheet.cell(column=1, row = next_table_index).value = compiler
            information_sheet.cell(column=1, row = next_table_index).fill = self.target3_color
            information_sheet.cell(column=2, row = next_table_index).value = release_info["__common__"]["configuration"]["releaseinfo"]["toolchain"][compiler]["version"]
            next_table_index += 1
        information_sheet.cell(column=1, row = next_table_index).value = "Middlewares:"
        next_table_index += 1
        for middleware in sorted(release_info["__common__"]["configuration"]["releaseinfo"]["middlewares"]):
            information_sheet.cell(column=1, row = next_table_index).value = middleware
            information_sheet.cell(column=1, row = next_table_index).fill = self.target3_color
            information_sheet.cell(column=2, row = next_table_index).value = release_info["__common__"]["configuration"]["releaseinfo"]["middlewares"][middleware]["version"]
            next_table_index += 1

        for cell in information_sheet._cells.values():
            cell.border = self.border
        for _range in information_sheet._merged_cells:
            c = _range.split(":")
            information_sheet[c[0]].border = self.border
            information_sheet[c[1]].border = self.border
        information_sheet.column_dimensions['B'].width = 25

        drivers_sheet.cell(column=1, row = 1).value = "Driver"
        drivers_sheet.cell(column=1, row = 1).font = self.target_font
        drivers_sheet.cell(column=1, row = 1).fill = self.target1_color
        soc_index = 2
        for index,soc in enumerate(sorted(config)):
            print "process drivers sheet: %s" % soc
            drivers_sheet.cell(column=soc_index, row=1).value = soc
            drivers_sheet.cell(column=soc_index, row=1).fill = self.target1_color
            drivers_sheet.cell(column=soc_index, row=1).font = self.target_font
            cur_column = re.sub(r'[0-9]', '', drivers_sheet.cell(column=soc_index, row=1).coordinate)
            drivers_sheet.column_dimensions[cur_column].width = len(soc)*1.3
            if not soc in soc_driver:
                continue
            if not len(soc_driver[soc]):
                continue
            for drv in soc_driver[soc]:
                count = 2
                while (drivers_sheet.cell(column=1, row=count).value != drv) and (drivers_sheet.cell(column=1, row=count).value):
                    count += 1
                drivers_sheet.cell(column=soc_index, row=count).value = "A"
                drivers_sheet.cell(column=soc_index, row=count).font = self.result_font
                if not drivers_sheet.cell(column=1, row=count).value:
                    drivers_sheet.cell(column=1, row=count).value = drv
                    drivers_sheet.cell(column=1, row=count).font = self.target_font
                    drivers_sheet.cell(column=1, row=count).fill = self.target3_color
            soc_index+=1

        max_size = len(drivers_sheet.cell(column=1, row=1).value)
        count = 2
        while drivers_sheet.cell(column=1, row=count).value:
            _value = len(drivers_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        drivers_sheet.column_dimensions['A'].width = max_size*1.3
        for _y in range(1,soc_index):
            for _x in range(1,count):
                drivers_sheet.cell(column=_y, row=_x).border = self.border
                drivers_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not drivers_sheet.cell(column=_y, row=_x).value:
                    drivers_sheet.cell(column=_y, row=_x).fill = self.na_color
                    drivers_sheet.cell(column=_y, row=_x).value = 'n/a'

        examples_status_sheet.cell(column=1, row = 1).value = "Demo_apps"
        examples_status_sheet.cell(column=1, row = 1).font = self.target_font
        examples_status_sheet.cell(column=1, row = 1).fill = self.target1_color
        examples_status_sheet.merge_cells(start_row=1,start_column=1,end_row=3,end_column=1)
        board_index = 2
        # fill data
        for index,board in enumerate(sorted(demos)):
            print "process demo_apps sheet: %s" % board
            if not demos[board]:
                continue
            if not demos[board]["soc"] in config:
                continue
            examples_status_sheet.cell(column=board_index, row=1).value = board
            examples_status_sheet.cell(column=board_index, row=1).fill = self.target1_color
            examples_status_sheet.cell(column=board_index, row=1).font = self.target_font
            for _index, compiler in enumerate(sorted(COMPILER)):
                examples_status_sheet.cell(column=board_index+_index*2, row=2).value = compiler
                examples_status_sheet.cell(column=board_index+_index*2, row=2).font = self.result_font
                examples_status_sheet.cell(column=board_index+_index*2, row=2).fill = self.target2_color
                examples_status_sheet.merge_cells(start_row=2,start_column=board_index+_index*2,end_row=2,end_column=board_index+_index*2+1)

                for __index, target in enumerate(TARGET):
                    examples_status_sheet.cell(column=board_index+_index*2+__index, row=3).value = target
                    examples_status_sheet.cell(column=board_index+_index*2+__index, row=3).font = self.result_font
                    examples_status_sheet.cell(column=board_index+_index*2+__index, row=3).fill = self.target2_color
                    cur_column = re.sub(r'[0-9]', '', examples_status_sheet.cell(column=board_index+_index*2+__index, row=2).coordinate)
                    examples_status_sheet.column_dimensions[cur_column].width = 4

                for demo in sorted(demos[board]):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (examples_status_sheet.cell(column=1, row=count).value != demo) and (examples_status_sheet.cell(column=1, row=count).value):
                        count += 1
                    if compiler in demos[board][demo]:
                        examples_status_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                        examples_status_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
                        if demos[board][demo][compiler] == "A":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif demos[board][demo][compiler] == "Ro":
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif demos[board][demo][compiler] == "Do":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                        else:
                            continue
                    else:
                        continue
                    if not examples_status_sheet.cell(column=1, row=count).value:
                        examples_status_sheet.cell(column=1, row=count).value = demo
                        examples_status_sheet.cell(column=1, row=count).font = self.target_font
                        examples_status_sheet.cell(column=1, row=count).fill = self.target3_color
            board_index += 10

        if sorted(examples_status_sheet._cells)[-1][1] == 1:
            next_table_index = sorted(examples_status_sheet._cells)[-1][0]
        else:
            next_table_index = sorted(examples_status_sheet._cells)[-1][0]+1
        examples_status_sheet.cell(column=1, row = next_table_index).value = "Driver_examples"
        examples_status_sheet.cell(column=1, row = next_table_index).font = self.target_font
        examples_status_sheet.cell(column=1, row = next_table_index).fill = self.target1_color
        examples_status_sheet.merge_cells(start_row=next_table_index,start_column=1,end_row=next_table_index,end_column=board_index - 1)

        next_table_index += 1
        for index,board in enumerate(sorted(examples)):
            print "process driver_examples sheet: %s" % board
            board_index = 2
            while (examples_status_sheet.cell(column=board_index, row=1).value != board) and (examples_status_sheet.cell(column=board_index, row=2).value):
                board_index += 10
            if not examples[board]:
                continue
            if not examples[board]["soc"] in config:
                continue
            for _index, compiler in enumerate(sorted(COMPILER)):
                for __index, demo in enumerate(sorted(examples[board])):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (examples_status_sheet.cell(column=1, row=count).value != demo) and (examples_status_sheet.cell(column=1, row=count).value):
                        count += 1
                    if compiler in examples[board][demo]:
                        examples_status_sheet.cell(column=board_index+_index*2, row=count).font = self.result_font
                        examples_status_sheet.cell(column=board_index+_index*2+1, row=count).font = self.result_font
                        if examples[board][demo][compiler] == "A":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif examples[board][demo][compiler] == "Ro":
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif examples[board][demo][compiler] == "Do":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                        else:
                            continue
                    else:
                        continue
                    if not examples_status_sheet.cell(column=1, row=count).value:
                        examples_status_sheet.cell(column=1, row=count).value = demo
                        examples_status_sheet.cell(column=1, row=count).font = self.target_font
                        examples_status_sheet.cell(column=1, row=count).fill = self.target3_color
            board_index += 10

        if sorted(examples_status_sheet._cells)[-1][1] == 1:
            next_table_index = sorted(examples_status_sheet._cells)[-1][0]
        else:
            next_table_index = sorted(examples_status_sheet._cells)[-1][0]+1
        examples_status_sheet.cell(column=1, row = next_table_index).value = "Usb"
        examples_status_sheet.cell(column=1, row = next_table_index).font = self.target_font
        examples_status_sheet.cell(column=1, row = next_table_index).fill = self.target1_color
        examples_status_sheet.merge_cells(start_row=next_table_index,start_column=1,end_row=next_table_index,end_column=board_index - 1)

        for index,board in enumerate(sorted(usbs)):
            print "process usb sheet: %s" % board
            board_index = 2
            while (examples_status_sheet.cell(column=board_index, row=1).value != board) and (examples_status_sheet.cell(column=board_index, row=2).value):
                board_index += 10
            if not usbs[board]:
                continue
            if not usbs[board]["soc"] in config:
                continue
            for _index, compiler in enumerate(sorted(COMPILER)):
                for __index, demo in enumerate(sorted(usbs[board])):
                    if demo == 'soc':
                        continue
                    count = 4
                    while (examples_status_sheet.cell(column=1, row=count).value != demo) and (examples_status_sheet.cell(column=1, row=count).value):
                        count += 1
                    examples_status_sheet.cell(column=board_index+_index, row=count).font = self.result_font
                    if compiler in usbs[board][demo]:
                        if usbs[board][demo][compiler] == "A":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif usbs[board][demo][compiler] == "Ro":
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2+1, row=count).fill = self.avail_color
                        elif usbs[board][demo][compiler] == "Do":
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).value = 'A'
                            examples_status_sheet.cell(column=board_index+_index*2, row=count).fill = self.avail_color
                        else:
                            continue
                    else:
                        continue
                    if not examples_status_sheet.cell(column=1, row=count).value:
                        examples_status_sheet.cell(column=1, row=count).value = demo
                        examples_status_sheet.cell(column=1, row=count).font = self.target_font
                        examples_status_sheet.cell(column=1, row=count).fill = self.target3_color

        max_size = len(examples_status_sheet.cell(column=1, row=1).value)
        count = 4
        while examples_status_sheet.cell(column=1, row=count).value:
            _value = len(examples_status_sheet.cell(column=1, row=count).value)
            if _value > max_size:
                max_size = _value
            count += 1
        examples_status_sheet.column_dimensions['A'].width = max_size*1.3

        board_index = 2
        while examples_status_sheet.cell(column=board_index, row=3).value:
            board_index += 1

        for _x in range(1,count):
            for _y in range(1,board_index):
                examples_status_sheet.cell(column=_y, row=_x).border = self.border
                examples_status_sheet.cell(column=_y, row=_x).alignment = self.alignment
                if not examples_status_sheet.cell(column=_y, row=_x).value and _x > 2 and _y > 1:
                    examples_status_sheet.cell(column=_y, row=_x).fill = self.na_color
                    examples_status_sheet.cell(column=_y, row=_x).value = 'n/a'
                    examples_status_sheet.cell(column=_y, row=_x).font = self.result_font
        start_merge = 2
        while examples_status_sheet.cell(column=start_merge, row=3).value:
            count = 1
            while ((not examples_status_sheet.cell(column=start_merge+count, row=1).value)
                and examples_status_sheet.cell(column=start_merge+count, row=3).value):
                count += 1
            if count > 1:
                examples_status_sheet.merge_cells(start_row=1,start_column=start_merge,end_row=1,end_column=start_merge+count-1)
                examples_status_sheet.cell(column=start_merge,row=1).font = self.target_font
                examples_status_sheet.cell(column=start_merge,row=1).fill = self.target1_color
            start_merge += count

        if not _report:
            _report = "checklist.xlsx"
        workbook.save(filename=_report)

if __name__ == "__main__":
    args = sys.argv
    try:
        __import__("openpyxl")
    except:
        import pip
        pip.main(["install", "openpyxl"])
    report = Reporter(args[1:])
